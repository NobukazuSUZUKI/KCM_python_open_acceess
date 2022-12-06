# Copyright © 2021 Nobukazu SUZUKI. All rights reserved.

# プログラムの内容：ビデオテープをデジタル化したmov(prores422hq)ファイルを対象にして、以下5点を実行する。
# 1. movからmp4(h264)をffmpegで作成する。
# 2. mp4をアクセスマスターとしてリネームする。
# 3. movとmp4のMD5を作成し、テキストファイルで保存する。
# 4. movとmp4を保存用のフォルダ構成にしてコピーする。
# 5. movとmp4のデータ容量を算出し、検査シート（エクセル）の特定セルに記入する。

# 事前インストールが必要となるソフトウェア2点。
# 1. ffmpeg (https://ffmpeg.org/)
# 2. openpyxl $pip install openpyxl

# TODO:ログファイルの書き出し
# TODO:ffmpegのコマンドを書き出さない設定

import datetime
import glob
import os
import shutil
import subprocess as sp
import hashlib
import csv
import math
import openpyxl #pip install openpyxl

# 計測の開始
print()
print('------------------------')
print("video_movツールを開始しました。")
dt_now = datetime.datetime.now() #現在時刻を取得
dt_now = dt_now.replace(microsecond = 0)
print(dt_now)
print()
print('-----------------------')

# 既にmp4が作成されているか確認する。
files_mov = glob.glob('./capture_HDD/*')
for file_mov_path in files_mov:
    file_mov = file_mov_path.replace(os.path.sep, '/') # パス区切りを'/'で統一し、winでもmacでも同一コードで動くようにする。
    file_mov_split = file_mov.split('/')
    file_name_mov = file_mov_split[-1]  # movファイル名を取得
    file_name_mov_split = file_name_mov.split('_')
    folder_name_video_id_title = file_name_mov_split[0] + '_' + file_name_mov_split[1] + '_' + file_name_mov_split[2]
    folder_name_video_id = file_name_mov_split[0] + '_' + file_name_mov_split[1]
    
    print('-----------------------')
    print('指定したmovファイルID: ' + folder_name_video_id_title)

    if os.path.exists('./KCM/' + folder_name_video_id_title):  #KCMフォルダ内にvideo_idフォルダがあるか確認する
        print('---mp4が既に有り。プログラムは実行しません。')
        print('-----------------------')
    else:
        print('---mp4が無し。プログラムを実行します。')
        dt_now = datetime.datetime.now() #現在時刻を取得
        start_time = dt_now.replace(microsecond = 0)
        print('開始時刻: ' + str(start_time))

        # KCMフォルダ内にvideo_#####_titleフォルダを作成し、movをコピーする。
        os.mkdir('./KCM/' + folder_name_video_id_title)
        source_mov = './capture_HDD/' + file_name_mov
        destination_mov = './KCM/' + folder_name_video_id_title
        shutil.copy2(source_mov, destination_mov)

        # mp4を作成する。ffmpeg
        input_file = source_mov
        output_file = destination_mov + '/' + file_name_mov + '.mp4'
        cmd_list = ['ffmpeg', '-i', input_file, '-c:v libx264', '-vf "yadif,scale=720:480"', '-pix_fmt yuv420p', '-b:v 5000k', '-ar 48000', '-ab 192k', output_file]
        cmd = ' '.join(cmd_list)
        sp.call(cmd, shell=True)

        # ffmpegコマンド例(MoMA)
        # ffmpeg -i [path to input file] -c:v prores -profile:v 3 -vf "yadif,scale=640:480" -pix_fmt yuv420p -c:a copy [path to output file]

        # mp4をリネームする
        before_fps = "720x486i"
        after_fps = "720x480p"
        before_word_codec = "prores422hq"
        after_word_codec = "h264"
        before_word_master = "preservationmaster.mov"
        after_word_master = "accessmaster"

        rename_before = output_file
        rename_fps = rename_before.replace(before_fps, after_fps)
        rename_codec = rename_fps.replace(before_word_codec, after_word_codec)
        rename_after = rename_codec.replace(before_word_master, after_word_master)
        os.rename(rename_before, rename_after)

        # MD5を書き出す。mov
        with open(source_mov, "rb") as open_file_MOV:  # ファイルをチャンクで開く_mov
            file_hash_MOV = hashlib.md5()
            while chunk_mov := open_file_MOV.read(8192):  # 65536 or 32768などで試しても良い
                file_hash_MOV.update(chunk_mov)
                hash_md5_MOV = file_hash_MOV.hexdigest()
                text_md5_MOV = hash_md5_MOV + ' *' + file_name_mov

        with open(destination_mov + '/' + file_name_mov + '.md5', "w") as file_md5_mov:  # md5テキストを作成_mov
            file_md5_mov.write(text_md5_MOV)

        # MD5を書き出す。mp4
        source_mp4 = rename_after
        file_mp4_split = source_mp4.split('/')  # macでもwinでも'/'で良い。output_fileとdestination_movで'/'と記述しているため。
        file_name_mp4 = file_mp4_split[3]  # mp4ファイル名を取得

        with open(source_mp4, "rb") as open_file_mp4:  #ファイルをチャンクで開く_mp4
            file_hash_mp4 = hashlib.md5()
            while chunk_mp4 := open_file_mp4.read(8192):  # 128 or 32768などで試しても良い。128 * 64 = 8192
                file_hash_mp4.update(chunk_mp4)
                hash_md5_mp4 = file_hash_mp4.hexdigest()
                text_md5_mp4 = hash_md5_mp4 + ' *' + file_name_mp4
        with open(source_mp4 + '.md5', "w") as file_md5_mp4:
            file_md5_mp4.write(text_md5_mp4)

        # データボリューム（容量）をGBに切り上げてcsvファイルに書き出す。
        volume_mov = os.path.getsize(source_mov)
        volume_mp4 = os.path.getsize(source_mp4)
        volume_mov_GB = volume_mov / 1024 ** 3
        volume_mp4_GB = volume_mp4 / 1024 ** 3
        volume_mov_GB_only = math.ceil(volume_mov_GB)  # 小数点切り上げ
        volume_mp4_GB_only = math.ceil(volume_mp4_GB)  # 小数点切り上げ

        #IDと比較し、エクセル内に容量が既に書かれているか確認する。
        desktop_dir = os.path.expanduser('~/Desktop') #デスクトップまでのディレクトリを取得（win,mac共通コード）
        wb = openpyxl.load_workbook(desktop_dir + '/' + 'videotape_inspection_sheet.xlsx') #エクセルを指定する。

        #IDとエクセル列を比較する。
        sheet = wb['Sheet1'] #シートを取得
        for row_ID in sheet.iter_cols(): #エクセルのIDセル列を上から下に順次検索する。
            for cell_ID in row_ID:
                if cell_ID.col_idx == 1: #エクセルのA列を指定する。
                    if cell_ID.value == folder_name_video_id: #エクセルA列のIDと、作成中のIDと一致するか確認する。
                        if not sheet.cell(row=cell_ID.row, column=30).value == None: #ID_titleセルと同一行の容量記入欄[AD列](30列目)が空欄かどうか確認する。
                            print('容量書き出し済み。')
                            print()
                        else:
                            # 容量をエクセルに書き込む
                            sheet.cell(row=cell_ID.row, column=30).value = volume_mov_GB_only
                            sheet.cell(row=cell_ID.row, column=31).value = volume_mp4_GB_only
                            wb.save(desktop_dir + '/' + 'videotape_inspection_sheet.xlsx') #エクセルを保存する。
                            wb.close()

        # 終了
        print("(MP4 & MD5 & 容量データを作成しました!")
        dt_now = datetime.datetime.now()  # 現在時刻を取得
        end_time = dt_now.replace(microsecond = 0)
        print('終了時刻：' + str(end_time))
        work_time = end_time - start_time
        print('作業時間：' + str(work_time))
        print('-----------------------')
        print()
